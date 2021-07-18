import openpyxl
from level_1.module import coins_module

wb = openpyxl.load_workbook("Score_level1.xlsx")
sheets = wb.sheetnames
sh1 = wb.active

def addit():
    d=1
    while d<100000 :
        
        eve = coins_module.Coin.num_coins_collected
        sh1.cell(row = d , column = 1 , value = eve)
        wb.save("Score_level1.xlsx")

        d=d+1

def checkit():

    max = sh1.cell(1,1).value
    d=2
    while d<100000:

        score = sh1.cell(row=d, column=1).value
        if int(score) > int(max):
            max=score    
        
        d=d+1
    
    return max